#!/usr/bin/env python

import argparse
import openpyxl
from datetime import datetime, timedelta
import re
import google_cal as cal

EVENT_SUMMARY = "Work shift"
EVENT_LOCATION = "Sveavägen 145, 113 46 Stockholm"
CALENDAR_ID = "oo3i1mcjbldcnjraltc6hi1ojk@group.calendar.google.com"
SHEETS = ["Personligt schema A (ro)", "Personligt schema B (ro)"]
NAME = "Olle W"
NAME_ROW = 2
TIME_ROW_START = 3
DATE_COLUMN = 4


def find_name_in_sheet(sheet, name: str):
    row = next(sheet.iter_rows(NAME_ROW, NAME_ROW))
    cell = next((c for c in row if c.value == name), None)
    return cell


def get_times(sheet, cell):
    times = [
        t[0] for t in sheet.iter_rows(
            TIME_ROW_START, 200, cell.column, cell.column, values_only=True)
    ]
    dates = [
        d[0] for d in sheet.iter_rows(
            TIME_ROW_START, 200, DATE_COLUMN, DATE_COLUMN, values_only=True)
    ]
    for start_date, time_string in zip(dates, times):
        if time_string == "Ledig":
            continue
        start_time_str, stop_time_str = re.findall(r'\d+', time_string)
        start_time = datetime.strptime(start_time_str, "%H")
        end_time = datetime.strptime(stop_time_str, "%H")
        start = datetime.combine(start_date, start_time.time())
        end = datetime.combine(start_date, end_time.time())
        if end < start:
            end = end + timedelta(days=+1)
        cal.add_event(start,
                      end,
                      summary=EVENT_SUMMARY,
                      location=EVENT_LOCATION,
                      cal_id=CALENDAR_ID)
        print(
            f"{start.strftime('%A %Y-%m-%d %H:%M')} - {end.strftime('%A %Y-%m-%d %H:%M')}"
        )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    args = parser.parse_args()
    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)
    ws = None
    for s in SHEETS:
        sheet = wb[s]
        cell = find_name_in_sheet(sheet, NAME)
        if cell:
            ws = sheet
            break
    get_times(ws, cell)


if __name__ == "__main__":
    main()
